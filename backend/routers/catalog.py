"""Catalog management endpoints."""
import csv
import io
import secrets
import tempfile
from pathlib import Path
from fastapi import APIRouter, HTTPException, UploadFile, File, Query
from pydantic import BaseModel
from openpyxl import load_workbook

from backend.models import CatalogEntry, CatalogUploadResponse, CatalogStatsResponse, ReferenceDocumentResponse
from procurement.services.database import (
    get_catalog_entries,
    get_catalog_skus,
    insert_catalog_entries,
    delete_catalog_entry,
    get_catalog_stats,
    batch_adjust_catalog_prices,
    save_reference_document,
    get_reference_documents,
    delete_reference_document,
)


class BatchAdjustPricesRequest(BaseModel):
    pct: float
    brand: str = None
    category: str = None

router = APIRouter(prefix="/api/catalog", tags=["catalog"])


@router.get("/stats", response_model=CatalogStatsResponse)
async def catalog_stats():
    """Get catalog statistics."""
    try:
        stats = get_catalog_stats()
        return CatalogStatsResponse(**stats)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/skus")
async def list_catalog_skus():
    """Get all SKUs in the catalog."""
    try:
        skus = get_catalog_skus()
        return skus
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/")
async def list_catalog_entries(
    search: str = Query(None),
    brand: str = Query(None),
    category: str = Query(None),
    limit: int = Query(500, ge=1, le=5000),
):
    """Search and list catalog entries."""
    try:
        entries = get_catalog_entries(search=search, brand=brand, category=category, limit=limit)
        return entries
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/upload", response_model=CatalogUploadResponse)
async def upload_catalog(file: UploadFile = File(...)):
    """Upload catalog entries from CSV or Excel file.

    Expected columns: sku (required), item_description, brand, base_price, min_price, max_price, currency, category
    """
    try:
        content = await file.read()
        entries = []
        errors = []

        # Determine file type
        if file.filename.endswith('.csv'):
            # Parse CSV
            text_content = content.decode('utf-8')
            reader = csv.DictReader(io.StringIO(text_content))
            for row_num, row in enumerate(reader, start=2):  # start at 2 for header
                if not row.get('sku'):
                    errors.append(f"Row {row_num}: Missing SKU")
                    continue
                try:
                    entry = {
                        'sku': row['sku'].strip().upper(),
                        'item_description': row.get('item_description', '').strip() or None,
                        'brand': row.get('brand', '').strip() or None,
                        'base_price': float(row['base_price']) if row.get('base_price') else None,
                        'min_price': float(row['min_price']) if row.get('min_price') else None,
                        'max_price': float(row['max_price']) if row.get('max_price') else None,
                        'currency': row.get('currency', 'USD').strip() or 'USD',
                        'category': row.get('category', '').strip() or None,
                    }
                    entries.append(entry)
                except ValueError as e:
                    errors.append(f"Row {row_num}: Invalid data - {str(e)}")

        elif file.filename.endswith(('.xlsx', '.xls')):
            # Parse Excel
            workbook = load_workbook(io.BytesIO(content))
            sheet = workbook.active
            headers = {}
            for col_num, cell in enumerate(sheet[1], start=1):
                if cell.value:
                    headers[cell.value.lower()] = col_num

            required_cols = ['sku']
            optional_cols = ['item_description', 'brand', 'base_price', 'min_price', 'max_price', 'currency', 'category']

            if not any(h in headers for h in required_cols):
                raise ValueError("Missing required column: sku")

            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
                sku_col = headers.get('sku')
                sku_cell = row[sku_col - 1] if sku_col else None
                sku = (sku_cell.value or '').strip().upper() if sku_cell else ''

                if not sku:
                    errors.append(f"Row {row_num}: Missing SKU")
                    continue

                try:
                    entry = {'sku': sku}
                    for col_name in optional_cols:
                        col_num = headers.get(col_name)
                        if col_num:
                            val = row[col_num - 1].value if col_num <= len(row) else None
                            if col_name in ('base_price', 'min_price', 'max_price'):
                                entry[col_name] = float(val) if val else None
                            else:
                                entry[col_name] = (val or '').strip() if isinstance(val, str) else str(val).strip() if val else None
                                if entry[col_name] == '':
                                    entry[col_name] = None
                    entries.append(entry)
                except ValueError as e:
                    errors.append(f"Row {row_num}: Invalid data - {str(e)}")

        else:
            raise ValueError("File must be CSV or Excel format (.csv, .xlsx, .xls)")

        # Insert entries
        inserted_count = insert_catalog_entries(entries)
        return CatalogUploadResponse(inserted_count=inserted_count, errors=errors)

    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.delete("/{entry_id}")
async def delete_catalog_entry_endpoint(entry_id: int):
    """Soft-delete a catalog entry."""
    try:
        success = delete_catalog_entry(entry_id)
        if not success:
            raise HTTPException(status_code=404, detail="Catalog entry not found")
        return {"status": "deleted", "id": entry_id}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/batch-adjust-prices")
async def batch_adjust_prices_endpoint(request: BatchAdjustPricesRequest):
    """Batch adjust catalog prices by percentage."""
    try:
        updated = batch_adjust_catalog_prices(
            pct=request.pct,
            brand=request.brand,
            category=request.category
        )
        return {"updated": updated}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/upload-reference-pdf", response_model=ReferenceDocumentResponse)
async def upload_reference_pdf(file: UploadFile = File(...)):
    """Upload a reference PDF catalog for fallback price lookups via RAG."""
    if not file.filename or not file.filename.lower().endswith('.pdf'):
        raise HTTPException(status_code=400, detail="Only PDF files are accepted")

    try:
        from procurement.services.llm_service import get_h2ogpte_client

        content = await file.read()
        client = get_h2ogpte_client()

        # Save to temp file for upload
        suffix = Path(file.filename).suffix
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp.write(content)
            tmp_path = tmp.name

        try:
            # Upload to H2OGPTE
            with open(tmp_path, 'rb') as f:
                upload_id = client.upload(file.filename, f)

            # Create persistent collection
            collection_name = f"ref_catalog_{secrets.token_hex(4)}"
            collection_id = client.create_collection(
                name=collection_name,
                description=f"Reference PDF catalog: {file.filename}",
            )
            client.ingest_uploads(collection_id, [upload_id], timeout=120)

            # Save to database
            disk_filename = f"{secrets.token_hex(8)}{suffix}"
            doc_id = save_reference_document(disk_filename, file.filename, collection_id)

            doc = {
                'id': doc_id,
                'filename': disk_filename,
                'original_name': file.filename,
                'collection_id': collection_id,
            }
            return ReferenceDocumentResponse(**doc)
        finally:
            Path(tmp_path).unlink(missing_ok=True)

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/reference-docs", response_model=list[ReferenceDocumentResponse])
async def list_reference_docs():
    """List all uploaded reference PDF catalogs."""
    try:
        docs = get_reference_documents()
        return [ReferenceDocumentResponse(**d) for d in docs]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/reference-docs/{doc_id}")
async def delete_reference_doc(doc_id: int):
    """Delete a reference PDF and its H2OGPTE collection."""
    try:
        doc = delete_reference_document(doc_id)
        if not doc:
            raise HTTPException(status_code=404, detail="Reference document not found")

        # Clean up H2OGPTE collection
        try:
            from procurement.services.llm_service import get_h2ogpte_client
            client = get_h2ogpte_client()
            client.delete_collections([doc['collection_id']])
        except Exception:
            pass  # Best-effort cleanup

        return {"status": "deleted", "id": doc_id}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
